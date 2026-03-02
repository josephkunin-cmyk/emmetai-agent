"""
Banyan Communications LLC — Emmet AI Phone Agent
Twilio + Anthropic Claude voice agent for Amish homesteads and rural communities.
"""

import os
import json
import logging
import sqlite3
import base64
import hashlib
import hmac
import re
import time
import uuid
import requests
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, List, Dict
from urllib.error import HTTPError
from urllib.request import Request, urlopen
from flask import Flask, request, Response
from twilio.rest import Client as TwilioClient
from twilio.request_validator import RequestValidator
from twilio.twiml.voice_response import VoiceResponse, Gather
from anthropic import Anthropic
from dotenv import load_dotenv
from zoneinfo import ZoneInfo

load_dotenv()

# ── Logging ──────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)
logger = logging.getLogger("emmet")


def env_int(name: str, default: int) -> int:
    raw = os.environ.get(name, str(default)).strip()
    try:
        return int(raw)
    except ValueError:
        return default


def env_text(name: str, default: str) -> str:
    value = os.environ.get(name, default)
    return (value or default).strip()


def env_bool(name: str, default: bool = False) -> bool:
    raw = os.environ.get(name, str(default)).strip().lower()
    return raw in {"1", "true", "yes", "on"}


def env_int_alias(primary: str, fallback: str, default: int) -> int:
    if os.environ.get(primary, "").strip():
        return env_int(primary, default)
    if os.environ.get(fallback, "").strip():
        return env_int(fallback, default)
    return default


BUSINESS_TIMEZONE = env_text("BUSINESS_TIMEZONE", "America/New_York")
FREE_DAILY_QUERIES = env_int_alias("FREE_DAILY_QUERIES", "FREE_QUERIES_PER_DAY", 5)
PAID_DAILY_QUERIES = env_int("PAID_DAILY_QUERIES", 4)
DB_PATH = env_text("DB_PATH", "./hotline_usage.db")
UPGRADE_MESSAGE = env_text(
    "UPGRADE_MESSAGE",
    (
        "You've used all your free questions for today. "
        "To ask four more questions, paid access is required. "
        "Text me back to unlock them, or call back tomorrow."
    ),
)
SERVICE_SCOPE_MESSAGE = env_text(
    "SERVICE_SCOPE_MESSAGE",
    (
        "This hotline is for agriculture, equestrian, homestead, and practical rural-life questions. "
        "Please ask a question in that scope."
    ),
)
SERVICE_GREETING = env_text(
    "SERVICE_GREETING",
    (
        "This is Emmet AI. I can answer your questions about farming and homesteading. Ask away."
    ),
)
VOICE_NAME = env_text("VOICE_NAME", "Polly.Joanna")
ANTHROPIC_MODEL = env_text("ANTHROPIC_MODEL", "claude-haiku-4-5-20251001")
APP_VERSION = env_text("APP_VERSION", "unknown")
TWILIO_ACCOUNT_SID = env_text("TWILIO_ACCOUNT_SID", "")
TWILIO_AUTH_TOKEN = env_text("TWILIO_AUTH_TOKEN", "")
TWILIO_MESSAGING_FROM = env_text("TWILIO_MESSAGING_FROM", "")
TWILIO_VALIDATE_SIGNATURE = env_bool("TWILIO_VALIDATE_SIGNATURE", False)
SERVICE_PHONE_DISPLAY = env_text("SERVICE_PHONE_DISPLAY", "+17179225968")
SQUARE_ACCESS_TOKEN = env_text("SQUARE_ACCESS_TOKEN", "")
SQUARE_LOCATION_ID = env_text("SQUARE_LOCATION_ID", "")
SQUARE_ENVIRONMENT = env_text("SQUARE_ENVIRONMENT", "production").lower()
SQUARE_API_VERSION = env_text("SQUARE_API_VERSION", "2025-10-16")
SQUARE_CURRENCY = env_text("SQUARE_CURRENCY", "USD").upper()
SQUARE_DAILY_UNLOCK_CENTS = env_int("SQUARE_DAILY_UNLOCK_CENTS", 500)
SQUARE_WEBHOOK_SIGNATURE_KEY = env_text("SQUARE_WEBHOOK_SIGNATURE_KEY", "")
PUBLIC_BASE_URL = env_text("PUBLIC_BASE_URL", "https://emmetai-agent.onrender.com").rstrip("/")

EMERGENCY_KEYWORDS = {
    "heart attack", "stroke", "can't breathe", "not breathing", "seizure",
    "overdose", "suicide", "kill myself", "self harm", "bleeding badly",
    "house fire", "barn fire", "emergency",
}

DISALLOWED_CONTENT_KEYWORDS = {
    "porn", "nude", "sex video", "erotic", "explicit",
    "make a bomb", "build a bomb", "buy cocaine", "meth recipe", "credit card fraud",
    "hack account", "phishing", "steal password", "weapon for attack",
}

OFF_TOPIC_KEYWORDS = {
    "celebrity", "movie review", "sports betting", "crypto day trading",
    "video game", "dating advice", "instagram growth", "tiktok strategy",
    "celebrity gossip", "hollywood",
}


class UsageStore:
    """SQLite-backed daily query counters keyed by caller and local date."""

    def __init__(self, db_path: str) -> None:
        self.db_path = db_path
        self._init_schema()

    def _connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn

    def _init_schema(self) -> None:
        with self._connect() as conn:
            conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS daily_usage (
                    caller_phone TEXT NOT NULL,
                    usage_date TEXT NOT NULL,
                    query_count INTEGER NOT NULL DEFAULT 0,
                    updated_at TEXT NOT NULL,
                    PRIMARY KEY (caller_phone, usage_date)
                );

                CREATE TABLE IF NOT EXISTS caller_profiles (
                    caller_phone TEXT PRIMARY KEY,
                    caller_name TEXT,
                    updated_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS paid_daily_access (
                    caller_phone TEXT NOT NULL,
                    usage_date TEXT NOT NULL,
                    status TEXT NOT NULL,
                    source TEXT,
                    updated_at TEXT NOT NULL,
                    PRIMARY KEY (caller_phone, usage_date)
                );

                CREATE TABLE IF NOT EXISTS square_checkout_links (
                    order_id TEXT PRIMARY KEY,
                    payment_link_id TEXT,
                    payment_link_url TEXT,
                    caller_phone TEXT NOT NULL,
                    usage_date TEXT NOT NULL,
                    amount_cents INTEGER NOT NULL,
                    paid_at TEXT,
                    created_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS call_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    call_sid TEXT NOT NULL,
                    caller_phone TEXT NOT NULL,
                    caller_name TEXT,
                    turn_number INTEGER NOT NULL DEFAULT 1,
                    user_message TEXT NOT NULL,
                    assistant_message TEXT NOT NULL,
                    latency_ms INTEGER,
                    usage_date TEXT NOT NULL,
                    created_at TEXT NOT NULL
                );
                CREATE INDEX IF NOT EXISTS idx_call_logs_sid ON call_logs(call_sid);
                CREATE INDEX IF NOT EXISTS idx_call_logs_date ON call_logs(usage_date);

                CREATE TABLE IF NOT EXISTS knowledge_updates (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    entry_type TEXT NOT NULL,
                    question TEXT,
                    answer TEXT,
                    announcement TEXT,
                    active INTEGER NOT NULL DEFAULT 1,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS marketplace_listings (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    category TEXT NOT NULL,
                    title TEXT NOT NULL,
                    description TEXT NOT NULL,
                    price TEXT,
                    location TEXT,
                    contact TEXT,
                    status TEXT NOT NULL DEFAULT 'available',
                    listed_by TEXT,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                );
                CREATE INDEX IF NOT EXISTS idx_market_category ON marketplace_listings(category);
                CREATE INDEX IF NOT EXISTS idx_market_status ON marketplace_listings(status);
                """
            )
            conn.commit()

    @staticmethod
    def _now_iso() -> str:
        return datetime.utcnow().isoformat(timespec="seconds")

    def get_count(self, caller_phone: str, usage_date: str) -> int:
        with self._connect() as conn:
            row = conn.execute(
                """
                SELECT query_count
                FROM daily_usage
                WHERE caller_phone = ? AND usage_date = ?
                """,
                (caller_phone, usage_date),
            ).fetchone()
        return int(row["query_count"]) if row else 0

    def increment(self, caller_phone: str, usage_date: str) -> int:
        now_iso = self._now_iso()
        with self._connect() as conn:
            conn.execute(
                """
                INSERT INTO daily_usage (caller_phone, usage_date, query_count, updated_at)
                VALUES (?, ?, 1, ?)
                ON CONFLICT(caller_phone, usage_date)
                DO UPDATE SET
                    query_count = query_count + 1,
                    updated_at = excluded.updated_at
                """,
                (caller_phone, usage_date, now_iso),
            )
            row = conn.execute(
                """
                SELECT query_count
                FROM daily_usage
                WHERE caller_phone = ? AND usage_date = ?
                """,
                (caller_phone, usage_date),
            ).fetchone()
            conn.commit()
        return int(row["query_count"]) if row else 1

    def get_caller_name(self, caller_phone: str) -> Optional[str]:
        with self._connect() as conn:
            row = conn.execute(
                """
                SELECT caller_name
                FROM caller_profiles
                WHERE caller_phone = ?
                """,
                (caller_phone,),
            ).fetchone()
        if not row:
            return None
        return (row["caller_name"] or "").strip() or None

    def upsert_caller_name(self, caller_phone: str, caller_name: str) -> None:
        now_iso = self._now_iso()
        with self._connect() as conn:
            conn.execute(
                """
                INSERT INTO caller_profiles (caller_phone, caller_name, updated_at)
                VALUES (?, ?, ?)
                ON CONFLICT(caller_phone)
                DO UPDATE SET
                    caller_name = excluded.caller_name,
                    updated_at = excluded.updated_at
                """,
                (caller_phone, caller_name, now_iso),
            )
            conn.commit()

    def has_paid_access(self, caller_phone: str, usage_date: str) -> bool:
        with self._connect() as conn:
            row = conn.execute(
                """
                SELECT status
                FROM paid_daily_access
                WHERE caller_phone = ? AND usage_date = ?
                """,
                (caller_phone, usage_date),
            ).fetchone()
        return bool(row and (row["status"] or "").lower() == "paid")

    def set_paid_access(self, caller_phone: str, usage_date: str, source: str) -> None:
        now_iso = self._now_iso()
        with self._connect() as conn:
            conn.execute(
                """
                INSERT INTO paid_daily_access (caller_phone, usage_date, status, source, updated_at)
                VALUES (?, ?, 'paid', ?, ?)
                ON CONFLICT(caller_phone, usage_date)
                DO UPDATE SET
                    status = 'paid',
                    source = excluded.source,
                    updated_at = excluded.updated_at
                """,
                (caller_phone, usage_date, source, now_iso),
            )
            conn.commit()

    def record_square_checkout(
        self,
        order_id: str,
        payment_link_id: str,
        payment_link_url: str,
        caller_phone: str,
        usage_date: str,
        amount_cents: int,
    ) -> None:
        now_iso = self._now_iso()
        with self._connect() as conn:
            conn.execute(
                """
                INSERT OR REPLACE INTO square_checkout_links (
                    order_id, payment_link_id, payment_link_url, caller_phone,
                    usage_date, amount_cents, paid_at, created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, NULL, ?)
                """,
                (
                    order_id,
                    payment_link_id,
                    payment_link_url,
                    caller_phone,
                    usage_date,
                    amount_cents,
                    now_iso,
                ),
            )
            conn.commit()

    def get_checkout_by_order(self, order_id: str) -> Optional[dict]:
        with self._connect() as conn:
            row = conn.execute(
                """
                SELECT order_id, caller_phone, usage_date, paid_at
                FROM square_checkout_links
                WHERE order_id = ?
                """,
                (order_id,),
            ).fetchone()
        return dict(row) if row else None

    def log_turn(
        self,
        call_sid: str,
        caller_phone: str,
        caller_name: str,
        user_message: str,
        assistant_message: str,
        latency_ms: int,
    ) -> None:
        """Write a single Q&A turn to the call_logs table."""
        now_iso = self._now_iso()
        usage_date = now_iso[:10]
        with self._connect() as conn:
            turn_number = (conn.execute(
                "SELECT COUNT(*) FROM call_logs WHERE call_sid = ?", (call_sid,)
            ).fetchone()[0] or 0) + 1
            conn.execute(
                """
                INSERT INTO call_logs
                    (call_sid, caller_phone, caller_name, turn_number,
                     user_message, assistant_message, latency_ms, usage_date, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (call_sid, caller_phone, caller_name or "", turn_number,
                 user_message, assistant_message, latency_ms, usage_date, now_iso),
            )
            conn.commit()

    def mark_checkout_paid(self, order_id: str) -> None:
        now_iso = self._now_iso()
        with self._connect() as conn:
            conn.execute(
                """
                UPDATE square_checkout_links
                SET paid_at = ?
                WHERE order_id = ?
                """,
                (now_iso, order_id),
            )
            conn.commit()

# ── App Setup ────────────────────────────────────────────────────────────
app = Flask(__name__)
anthropic_client = None
usage_store = UsageStore(DB_PATH)

# In-memory conversation store keyed by Twilio CallSid
conversations = {}
call_metadata = {}


def business_today_iso() -> str:
    try:
        now_local = datetime.now(ZoneInfo(BUSINESS_TIMEZONE))
    except Exception:  # fallback for invalid timezone configuration
        now_local = datetime.utcnow()
    return now_local.date().isoformat()


def caller_identity(call_sid: str, caller_raw: str) -> str:
    caller = (caller_raw or "").strip()
    if caller and caller.lower() != "unknown":
        return caller
    if call_sid:
        return f"call:{call_sid}"
    return "unknown"


def guardrail_response(user_text: str) -> Optional[str]:
    text = (user_text or "").strip().lower()
    if not text:
        return None

    if any(keyword in text for keyword in EMERGENCY_KEYWORDS):
        return (
            "If this is an emergency, please hang up and call 911 right now. "
            "For immediate mental health crisis help, call or text 988."
        )

    if any(keyword in text for keyword in DISALLOWED_CONTENT_KEYWORDS):
        return (
            "I can't help with that. This hotline stays PG-13 and does not support harmful, sexual, or illegal requests. "
            "I can help with safe agriculture, horse care, and practical rural questions."
        )

    if any(keyword in text for keyword in OFF_TOPIC_KEYWORDS):
        return SERVICE_SCOPE_MESSAGE

    return None


def extract_first_name(raw_text: str, fallback: str = "friend") -> str:
    text = (raw_text or "").strip()
    if not text:
        return fallback
    lowered = text.lower()
    for prefix in ("my name is", "this is", "i am", "i'm", "its", "it is"):
        if lowered.startswith(prefix):
            text = text[len(prefix):].strip()
            break
    clean = re.sub(r"[^A-Za-z\s\-']", " ", text)
    parts = [p for p in clean.split() if p]
    if not parts:
        return fallback
    first = parts[0][:30]
    return first.capitalize()


def is_e164_phone(value: str) -> bool:
    return bool(re.fullmatch(r"\+\d{8,15}", (value or "").strip()))


def get_twilio_client() -> Optional[TwilioClient]:
    if not TWILIO_ACCOUNT_SID or not TWILIO_AUTH_TOKEN:
        return None
    try:
        return TwilioClient(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)
    except Exception as exc:  # noqa: BLE001
        logger.error(f"Failed to initialize Twilio client: {exc}")
        return None


def enforce_twilio_signature() -> Optional[Response]:
    """Validate Twilio request signatures when enabled."""
    if request.method != "POST":
        return None
    if not TWILIO_VALIDATE_SIGNATURE:
        return None
    if not TWILIO_AUTH_TOKEN:
        logger.warning("TWILIO_VALIDATE_SIGNATURE=true but TWILIO_AUTH_TOKEN is missing")
        return Response("twilio auth token missing", status=500)
    signature = request.headers.get("X-Twilio-Signature", "").strip()
    if not signature:
        return Response("missing twilio signature", status=403)
    validator = RequestValidator(TWILIO_AUTH_TOKEN)
    valid = validator.validate(
        request.url,
        request.form.to_dict(flat=True),
        signature,
    )
    if not valid:
        return Response("invalid twilio signature", status=403)
    return None


def send_payment_sms(caller_phone: str, payment_url: str) -> bool:
    if not is_e164_phone(caller_phone):
        return False
    if not TWILIO_MESSAGING_FROM:
        return False
    client = get_twilio_client()
    if client is None:
        return False
    text = (
        "Banyan Communications: you've hit today's free limit. "
        f"Use this secure link to unlock paid access now: {payment_url}"
    )
    try:
        client.messages.create(
            to=caller_phone,
            from_=TWILIO_MESSAGING_FROM,
            body=text,
        )
        return True
    except Exception as exc:  # noqa: BLE001
        logger.error(f"SMS send failed for {caller_phone}: {exc}")
        return False


def square_api_base_url() -> str:
    if SQUARE_ENVIRONMENT == "sandbox":
        return "https://connect.squareupsandbox.com"
    return "https://connect.squareup.com"


def create_square_payment_link(caller_phone: str, usage_date: str) -> Optional[str]:
    if not SQUARE_ACCESS_TOKEN or not SQUARE_LOCATION_ID:
        return None
    payload = {
        "idempotency_key": str(uuid.uuid4()),
        "quick_pay": {
            "name": "Emmet AI Daily Unlock",
            "price_money": {
                "amount": SQUARE_DAILY_UNLOCK_CENTS,
                "currency": SQUARE_CURRENCY,
            },
            "location_id": SQUARE_LOCATION_ID,
        },
        "description": f"Daily unlock for {caller_phone} on {usage_date}",
    }
    if PUBLIC_BASE_URL:
        payload["checkout_options"] = {"redirect_url": f"{PUBLIC_BASE_URL}/health"}
    body = json.dumps(payload).encode("utf-8")
    endpoint = f"{square_api_base_url()}/v2/online-checkout/payment-links"
    req = Request(endpoint, data=body, method="POST")
    req.add_header("Authorization", f"Bearer {SQUARE_ACCESS_TOKEN}")
    req.add_header("Square-Version", SQUARE_API_VERSION)
    req.add_header("Content-Type", "application/json")

    try:
        with urlopen(req, timeout=20) as response:
            data = json.loads(response.read().decode("utf-8"))
    except HTTPError as exc:
        details = exc.read().decode("utf-8", errors="ignore")
        logger.error(f"Square payment-link HTTP error: {exc.code} {details}")
        return None
    except Exception as exc:  # noqa: BLE001
        logger.error(f"Square payment-link failure: {exc}")
        return None

    payment_link = data.get("payment_link") or {}
    payment_link_url = (payment_link.get("url") or "").strip()
    payment_link_id = (payment_link.get("id") or "").strip()
    order_id = (payment_link.get("order_id") or "").strip()

    related_resources = data.get("related_resources") or {}
    if not order_id:
        orders = related_resources.get("orders") or []
        if orders:
            order_id = (orders[0].get("id") or "").strip()

    if payment_link_url and order_id:
        usage_store.record_square_checkout(
            order_id=order_id,
            payment_link_id=payment_link_id,
            payment_link_url=payment_link_url,
            caller_phone=caller_phone,
            usage_date=usage_date,
            amount_cents=SQUARE_DAILY_UNLOCK_CENTS,
        )
    return payment_link_url or None


def build_limit_message(caller_phone: str, usage_date: str) -> str:
    payment_url = create_square_payment_link(caller_phone, usage_date)
    if payment_url:
        if send_payment_sms(caller_phone, payment_url):
            return (
                f"{UPGRADE_MESSAGE} I just texted a secure payment link to this phone number. "
                "After payment, call back and I can keep helping today."
            )
        return (
            f"{UPGRADE_MESSAGE} A secure payment link is ready. "
            "Please contact Banyan Communications if you need help completing payment."
        )
    return UPGRADE_MESSAGE


def verify_square_webhook(payload: bytes) -> bool:
    if not SQUARE_WEBHOOK_SIGNATURE_KEY:
        return True
    signature = request.headers.get("x-square-hmacsha256-signature", "").strip()
    if not signature:
        return False
    signed_payload = request.url.encode("utf-8") + payload
    digest = hmac.new(
        SQUARE_WEBHOOK_SIGNATURE_KEY.encode("utf-8"),
        signed_payload,
        hashlib.sha256,
    ).digest()
    expected = base64.b64encode(digest).decode("utf-8")
    return hmac.compare_digest(expected, signature)


def apply_paid_event_from_square(payment: dict) -> None:
    status = (payment.get("status") or "").upper()
    if status != "COMPLETED":
        return
    order_id = (payment.get("order_id") or "").strip()
    if not order_id:
        return
    checkout = usage_store.get_checkout_by_order(order_id)
    if not checkout:
        return
    usage_store.mark_checkout_paid(order_id)
    usage_store.set_paid_access(
        checkout["caller_phone"],
        checkout["usage_date"],
        source="square_webhook",
    )
    logger.info(
        "Square payment completed for %s on %s",
        checkout["caller_phone"],
        checkout["usage_date"],
    )

# ── Knowledge Base ───────────────────────────────────────────────────────
KNOWLEDGE_BASE_PATH = os.path.join(os.path.dirname(__file__), "knowledge_base.json")

def load_knowledge_base():
    """Load knowledge base — DB overrides first, then falls back to JSON file."""
    try:
        with open(KNOWLEDGE_BASE_PATH, "r") as f:
            base = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        base = {"entries": [], "announcements": []}

    # Merge in live DB updates
    try:
        with usage_store._connect() as conn:
            rows = conn.execute(
                "SELECT * FROM knowledge_updates WHERE active=1 ORDER BY id"
            ).fetchall()
        for row in rows:
            if row["entry_type"] == "qa" and row["question"] and row["answer"]:
                base["entries"].append({"question": row["question"], "answer": row["answer"]})
            elif row["entry_type"] == "announcement" and row["announcement"]:
                base["announcements"].append(row["announcement"])
    except Exception:
        pass
    return base

def format_knowledge_for_prompt():
    """Format knowledge base entries for the system prompt."""
    kb = load_knowledge_base()
    sections = []

    if kb.get("announcements"):
        announcements = "\n".join(f"- {a}" for a in kb["announcements"])
        sections.append(f"CURRENT COMMUNITY ANNOUNCEMENTS:\n{announcements}")

    if kb.get("entries"):
        entries = "\n".join(
            f"Q: {e['question']}\nA: {e['answer']}"
            for e in kb["entries"]
        )
        sections.append(f"COMMUNITY KNOWLEDGE BASE:\n{entries}")

    return "\n\n".join(sections)


# ── System Prompt ────────────────────────────────────────────────────────
def format_marketplace_for_prompt() -> str:
    """Pull active marketplace listings into a prompt section."""
    try:
        with usage_store._connect() as conn:
            rows = conn.execute(
                """SELECT category, title, description, price, location, contact, status
                   FROM marketplace_listings ORDER BY category, status, id DESC"""
            ).fetchall()
        if not rows:
            return ""
        sections = {}
        for r in rows:
            cat = r["category"].title()
            sections.setdefault(cat, [])
            status_tag = "" if r["status"] == "available" else f" [{r['status'].upper()}]"
            parts = [f"{r['title']}{status_tag}: {r['description']}"]
            if r["price"]:
                parts.append(f"Price: {r['price']}")
            if r["location"]:
                parts.append(f"Location: {r['location']}")
            if r["contact"]:
                parts.append(f"Contact: {r['contact']}")
            sections[cat].append(" | ".join(parts))
        lines = []
        for cat, entries in sections.items():
            lines.append(f"{cat.upper()} LISTINGS:")
            lines.extend(f"  - {e}" for e in entries)
        return "LIVE MARKETPLACE LISTINGS (current buy/sell/trade listings in the community):\n" + "\n".join(lines)
    except Exception:
        return ""


# ── Farming Knowledge Database ──────────────────────────────────────────────
FARMING_KNOWLEDGE_DB = os.path.join(os.path.dirname(__file__), "farming_knowledge.db")

def initialize_farming_knowledge_db():
    """Initialize farming knowledge database with seed data."""
    try:
        # Only initialize if not already done
        if os.path.exists(FARMING_KNOWLEDGE_DB):
            return

        conn = sqlite3.connect(FARMING_KNOWLEDGE_DB)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Create tables
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS knowledge (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                category TEXT NOT NULL,
                topic TEXT,
                content TEXT NOT NULL,
                source TEXT DEFAULT 'seed-data',
                added_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
        """)

        # Seed data for farming knowledge base
        farming_knowledge_base = {
            "weather_&_seasons": [
                {
                    "topic": "Frost dates",
                    "region": "Lancaster County PA",
                    "content": "Last spring frost: ~May 10. First fall frost: ~October 1. Growing season: ~144 days."
                },
                {
                    "topic": "Planting calendar",
                    "month": "March",
                    "content": "Start seeds indoors: tomatoes, peppers, eggplant. Direct sow: peas, spinach, lettuce, kale."
                },
                {
                    "topic": "Planting calendar",
                    "month": "April",
                    "content": "Plant: asparagus, rhubarb, garlic. Direct sow: carrots, beets, beans, corn after last frost date."
                },
                {
                    "topic": "Planting calendar",
                    "month": "May",
                    "content": "Transplant: tomatoes, peppers, eggplant after last frost. Direct sow: beans, corn, squash, cucumbers."
                }
            ],
            "crops": [
                {
                    "crop": "Tomatoes",
                    "spacing": "24-36 inches apart",
                    "days_to_harvest": "60-85 days",
                    "sunlight": "6-8 hours minimum",
                    "tips": "Use cages or stakes. Prune suckers for better fruit. Rotate crops yearly."
                },
                {
                    "crop": "Corn",
                    "spacing": "8-12 inches apart, rows 30 inches",
                    "days_to_harvest": "60-100 days depending on variety",
                    "sunlight": "Full sun (6+ hours)",
                    "tips": "Plant in blocks for good pollination. Water deeply. Harvest when silks brown."
                },
                {
                    "crop": "Beans",
                    "spacing": "4-6 inches apart",
                    "days_to_harvest": "50-60 days",
                    "sunlight": "6+ hours",
                    "tips": "Don't soak seeds. Plant after last frost. Bush beans faster than pole."
                },
                {
                    "crop": "Potatoes",
                    "spacing": "12 inches apart, rows 3 feet",
                    "days_to_harvest": "70-120 days",
                    "sunlight": "6+ hours",
                    "tips": "Plant seed potatoes 2-4 inches deep. Hill soil around plants as they grow."
                }
            ],
            "livestock": [
                {
                    "animal": "Chickens",
                    "space_per_bird": "3-4 sq ft indoor, 8-10 sq ft outdoor",
                    "feed": "~0.25 lb per day (layer feed)",
                    "water": "Continuous access. ~1 cup per day.",
                    "tips": "Provide ventilation not drafts. Predator-proof coop essential. Expect 5-6 eggs/week per hen."
                },
                {
                    "animal": "Goats",
                    "space_per_animal": "200+ sq ft pasture per goat",
                    "feed": "2-3% of body weight daily (hay + grain)",
                    "water": "~1 gallon per day per 100 lbs",
                    "tips": "Goats are escape artists. Strong fencing needed. Can eat brush and weeds."
                },
                {
                    "animal": "Horses",
                    "space_per_animal": "1-2 acres per horse",
                    "feed": "1.5-2% of body weight daily (hay + grain)",
                    "water": "5-10 gallons per day",
                    "tips": "Require shelter, regular farrier care (8-10 weeks), dental care, vaccines."
                },
                {
                    "animal": "Pigs",
                    "space_per_animal": "50 sq ft per pig (minimum)",
                    "feed": "3-6 lbs per day depending on size/stage",
                    "water": "Continuous access, 1-2 gallons per day",
                    "tips": "Provide mud wallow or shade. Good for clearing land. 6-12 month grow-out."
                }
            ],
            "soil_care": [
                {
                    "practice": "Crop rotation",
                    "description": "Don't plant same crop family in same spot 2 years running",
                    "benefit": "Prevents disease buildup, improves soil",
                    "timeline": "Rotate every 1-2 years"
                },
                {
                    "practice": "Cover cropping",
                    "description": "Plant clover, alfalfa, rye in off-season",
                    "benefit": "Fixes nitrogen, prevents erosion, adds organic matter",
                    "timeline": "Plant fall, till in spring"
                },
                {
                    "practice": "Composting",
                    "description": "Kitchen scraps + yard waste → black gold",
                    "benefit": "Rich in nutrients. Improves soil structure. Saves money.",
                    "timeline": "3-12 months depending on method"
                }
            ],
            "food_preservation": [
                {
                    "method": "Canning",
                    "foods": "Tomatoes, salsa, jams, pickles, beans",
                    "shelf_life": "1-2 years",
                    "equipment": "Canner, jars, lids, pectin for jams",
                    "safety": "Follow USDA guidelines. Use pressure canner for low-acid foods."
                },
                {
                    "method": "Freezing",
                    "foods": "Vegetables, fruits, herbs, prepared meals",
                    "shelf_life": "6-12 months",
                    "equipment": "Freezer, freezer bags, vacuum sealer (optional)",
                    "tips": "Blanch vegetables first. Label with date. Store at 0°F."
                },
                {
                    "method": "Root cellar storage",
                    "foods": "Potatoes, onions, apples, squash, carrots",
                    "shelf_life": "2-6 months depending on crop",
                    "conditions": "Cool (32-50°F), dark, humid (90%+)",
                    "tips": "Store away from ethylene-producing fruits (apples)."
                },
                {
                    "method": "Dehydrating",
                    "foods": "Herbs, peppers, tomatoes, apples, jerky",
                    "shelf_life": "6-12 months",
                    "equipment": "Dehydrator or oven on low",
                    "tips": "Store in airtight containers. Use oxygen absorbers for long-term."
                }
            ],
            "tools_equipment": [
                {
                    "tool": "Hoe",
                    "use": "Weeding, breaking soil, making rows",
                    "types": "Standard, warren (pointed), warren/push combo"
                },
                {
                    "tool": "Tiller",
                    "use": "Breaking ground, prepping beds, mixing soil",
                    "types": "Front-tine (small), rear-tine (large), mini tillers"
                },
                {
                    "tool": "Shovel vs Spade",
                    "use": "Shovel: moving loose material. Spade: digging, edging.",
                    "types": "Long handle vs D-handle"
                }
            ]
        }

        # Populate with seed data
        now = datetime.utcnow().isoformat()
        for category, items in farming_knowledge_base.items():
            for item in items:
                content = json.dumps(item)
                topic = item.get("topic") or item.get("crop") or item.get("animal") or item.get("method") or item.get("tool") or item.get("practice")
                cursor.execute(
                    "INSERT INTO knowledge (category, topic, content, added_at, updated_at) VALUES (?,?,?,?,?)",
                    (category, topic, content, now, now)
                )

        conn.commit()
        conn.close()
        logger.info(f"✅ Farming knowledge database initialized at {FARMING_KNOWLEDGE_DB}")
    except Exception as e:
        logger.error(f"Error initializing farming knowledge DB: {e}")


def search_farming_knowledge(query: str, limit: int = 3) -> List[Dict]:
    """Search farming knowledge database for relevant information by keywords."""
    try:
        if not os.path.exists(FARMING_KNOWLEDGE_DB):
            return []

        conn = sqlite3.connect(FARMING_KNOWLEDGE_DB)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Extract keywords from query (min 3 chars to avoid noise)
        keywords = [w.lower() for w in query.split() if len(w) >= 3]
        if not keywords:
            return []

        # Build query searching for any keyword match
        # Start with topic matches (exact keywords are more important)
        topic_query = " OR ".join(["topic LIKE ?" for _ in keywords])
        topic_params = [f"%{kw}%" for kw in keywords]

        cursor.execute(f"""
            SELECT topic, content, category FROM knowledge
            WHERE {topic_query}
            LIMIT ?
        """, topic_params + [limit])

        results = []
        seen_topics = set()

        for row in cursor.fetchall():
            if row["topic"] in seen_topics:
                continue
            seen_topics.add(row["topic"])

            try:
                content_obj = json.loads(row["content"])
                results.append({
                    "topic": row["topic"],
                    "category": row["category"],
                    "content": content_obj
                })
            except json.JSONDecodeError:
                results.append({
                    "topic": row["topic"],
                    "category": row["category"],
                    "content": row["content"]
                })

            if len(results) >= limit:
                break

        # If we didn't find enough from topic matches, search content
        if len(results) < limit:
            content_query = " OR ".join(["content LIKE ?" for _ in keywords])
            content_params = [f"%{kw}%" for kw in keywords]

            cursor.execute(f"""
                SELECT topic, content, category FROM knowledge
                WHERE {content_query}
                LIMIT ?
            """, content_params + [limit - len(results)])

            for row in cursor.fetchall():
                if row["topic"] not in seen_topics:
                    seen_topics.add(row["topic"])
                    try:
                        content_obj = json.loads(row["content"])
                        results.append({
                            "topic": row["topic"],
                            "category": row["category"],
                            "content": content_obj
                        })
                    except json.JSONDecodeError:
                        results.append({
                            "topic": row["topic"],
                            "category": row["category"],
                            "content": row["content"]
                        })

        conn.close()
        return results
    except Exception as e:
        logger.error(f"Error searching farming knowledge: {e}")
        return []


def format_farming_knowledge_for_prompt(user_message: str) -> str:
    """Format farming knowledge search results for inclusion in system prompt."""
    try:
        results = search_farming_knowledge(user_message, limit=3)
        if not results:
            return ""

        lines = ["RELEVANT FARMING & HOMESTEADING KNOWLEDGE:"]
        for r in results:
            lines.append(f"\n{r['topic']} ({r['category']}):")
            content = r["content"]
            if isinstance(content, dict):
                # Format dict content nicely
                for key, val in content.items():
                    if key not in ["topic", "crop", "animal", "method", "tool", "practice"]:
                        lines.append(f"  {key.replace('_', ' ').title()}: {val}")
            else:
                lines.append(f"  {content}")

        return "\n".join(lines)
    except Exception as e:
        logger.error(f"Error formatting farming knowledge: {e}")
        return ""


def build_system_prompt():
    """Build the full system prompt including knowledge base."""
    base_prompt = """You are Emmet, a friendly and practical AI phone assistant for Banyan Communications.

Voice style:
- Warm, respectful, and plainspoken.
- Speak naturally for phone audio, no bullet points.
- Keep answers short: usually 2 to 4 sentences.
- Use practical advice first.

Scope:
- Agriculture, farming, gardening, livestock, horse and equestrian care.
- Homesteading and rural practical life: weather, tools, repairs, food preservation, measurements.
- PG-13 content only.

Guardrails:
- If a request is outside scope, politely decline and steer back to allowed topics.
- Refuse harmful, illegal, or explicit sexual content.
- For emergencies, tell caller to contact 911 immediately.
- For medical or veterinary high-risk situations, advise urgent professional help.
- Never shame the caller. Keep tone calm and respectful.

Identity:
- If asked your name, say: I'm Emmet, your AI assistant from Banyan Communications."""

    kb_text = format_knowledge_for_prompt()
    market_text = format_marketplace_for_prompt()
    extra = "\n\n".join(filter(None, [kb_text, market_text]))
    if extra:
        return f"{base_prompt}\n\n{extra}"
    return base_prompt


# ── Claude Integration ───────────────────────────────────────────────────
def get_conversation(call_sid):
    """Get or create conversation history for a call."""
    if call_sid not in conversations:
        conversations[call_sid] = []
    return conversations[call_sid]


def get_anthropic_client():
    """Lazily initialize the Anthropic client so boot never fails."""
    global anthropic_client
    if anthropic_client is not None:
        return anthropic_client

    api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        raise RuntimeError("ANTHROPIC_API_KEY is not set")

    anthropic_client = Anthropic(api_key=api_key)
    return anthropic_client


def ask_claude(call_sid: str, user_message: str) -> str:
    """Send message to Claude Haiku with prompt caching for low latency."""
    history = get_conversation(call_sid)
    history.append({"role": "user", "content": user_message})

    logger.info(f"[{call_sid[:8]}] Caller: {user_message}")

    t0 = time.monotonic()

    # Build system prompt with base knowledge + farming-specific knowledge for this question
    base_system = build_system_prompt()
    farming_knowledge = format_farming_knowledge_for_prompt(user_message)

    if farming_knowledge:
        system_prompt_text = f"{base_system}\n\n{farming_knowledge}"
    else:
        system_prompt_text = base_system

    # Use prompt caching on the system prompt — cuts repeat-call latency ~60%
    response = get_anthropic_client().messages.create(
        model=ANTHROPIC_MODEL,
        max_tokens=150,                        # phone answers are short
        system=[
            {
                "type": "text",
                "text": system_prompt_text,
                "cache_control": {"type": "ephemeral"},   # cache the prompt
            }
        ],
        messages=history,
    )

    latency_ms = int((time.monotonic() - t0) * 1000)
    assistant_message = response.content[0].text.strip()
    history.append({"role": "assistant", "content": assistant_message})

    logger.info(f"[{call_sid[:8]}] Emmet ({latency_ms}ms): {assistant_message}")

    # Write turn to call log
    caller_id = call_metadata.get(call_sid, {}).get("from", "unknown")
    caller_name = call_metadata.get(call_sid, {}).get("caller_name", "")
    usage_store.log_turn(call_sid, caller_id, caller_name, user_message, assistant_message, latency_ms)

    # Keep conversation history manageable (last 20 turns)
    if len(history) > 20:
        conversations[call_sid] = history[-20:]

    return assistant_message


# ── TwiML Response Builder ──────────────────────────────────────────────
def twiml_listen(
    text: str,
    action: str = "/gather",
    fallback_text: str = "I'm still here. Go ahead whenever you're ready.",
    redirect_url: str = "/voice?returning=true",
) -> str:
    """Build TwiML that speaks text and listens for a reply."""
    resp = VoiceResponse()
    gather = Gather(
        input="speech",
        action=action,
        method="POST",
        speech_timeout="2",       # 2s silence = end of turn (was "auto" ~3-4s)
        speech_model="phone_call",
        enhanced=True,
        language="en-US",
        profanity_filter=False,   # faster transcription
        action_on_empty_result=False,
    )
    gather.say(text, voice=VOICE_NAME, language="en-US")
    resp.append(gather)

    # Fallback if no speech detected
    resp.say(fallback_text, voice=VOICE_NAME, language="en-US")
    resp.redirect(redirect_url)
    return str(resp)


def twiml_say(text: str) -> str:
    """Build TwiML that speaks text and hangs up."""
    resp = VoiceResponse()
    resp.say(text, voice=VOICE_NAME, language="en-US")
    resp.hangup()
    return str(resp)


# ── Conversation Export to Excel ──────────────────────────────────────────
def export_call_to_spreadsheet(caller_phone: str, caller_name: str, usage_date: str, questions_asked: list, paid_status: bool) -> None:
    """Export a completed call to an Excel spreadsheet on the user's computer."""
    try:
        from openpyxl import load_workbook, Workbook

        # Determine export path: ~/Downloads/emmet-conversations.xlsx
        export_path = Path.home() / "Downloads" / "emmet-conversations.xlsx"
        export_path.parent.mkdir(parents=True, exist_ok=True)

        # Load or create workbook
        if export_path.exists():
            wb = load_workbook(export_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Conversations"
            # Add headers
            headers = ["Date", "Phone", "Name", "Questions Asked", "Paid Access", "Q1", "A1", "Q2", "A2", "Q3", "A3", "Q4", "A4", "Q5", "A5", "Q6", "A6", "Q7", "A7", "Q8", "A8"]
            ws.append(headers)

        # Prepare row data
        row_data = [
            datetime.utcnow().isoformat(),
            caller_phone,
            caller_name or "(not captured)",
            len(questions_asked),
            "Yes" if paid_status else "No"
        ]

        # Add Q&A pairs (up to 8 questions)
        for i in range(8):
            if i < len(questions_asked):
                q, a = questions_asked[i]
                row_data.append(q[:100])  # Truncate long questions
                row_data.append(a[:100])  # Truncate long answers
            else:
                row_data.append("")
                row_data.append("")

        # Append row to spreadsheet
        ws.append(row_data)

        # Save workbook
        wb.save(export_path)
        logger.info(f"✅ Call exported to {export_path}")
    except Exception as e:
        logger.error(f"Failed to export to spreadsheet: {e}")


# ── Routes ───────────────────────────────────────────────────────────────
@app.route("/voice", methods=["GET", "POST"])
def voice():
    """Handle incoming calls — greet the caller."""
    sig_err = enforce_twilio_signature()
    if sig_err:
        return sig_err

    call_sid = request.form.get("CallSid", "unknown")
    caller = request.form.get("From", "unknown")
    caller_id = caller_identity(call_sid, caller)
    returning = request.args.get("returning", "false")
    stage = request.args.get("stage", "").strip().lower()
    usage_date = business_today_iso()

    if returning == "true":
        if stage == "name":
            existing_name = usage_store.get_caller_name(caller_id)
            name_prompt = "Before we start, what's your first name?"
            if existing_name:
                name_prompt = (
                    f"Welcome back {existing_name}. "
                    "What name should I call you today? You can say same name."
                )
            return Response(
                twiml_listen(
                    name_prompt,
                    action="/intro-name",
                    fallback_text="I didn't catch your name. Please say your first name.",
                    redirect_url="/voice?returning=true&stage=name",
                ),
                mimetype="text/xml",
            )
        return Response(
            twiml_listen("Go ahead, I'm listening."),
            mimetype="text/xml"
        )

    # New call — clear old conversation
    conversations.pop(call_sid, None)

    # Track call metadata
    call_metadata[call_sid] = {
        "from": caller_id,
        "start": datetime.utcnow().isoformat(),
        "turns": 0
    }

    used_today = usage_store.get_count(caller_id, usage_date)
    has_paid = usage_store.has_paid_access(caller_id, usage_date)
    total_limit = FREE_DAILY_QUERIES + (PAID_DAILY_QUERIES if has_paid else 0)

    if used_today >= FREE_DAILY_QUERIES and not has_paid:
        limit_message = build_limit_message(caller_id, usage_date)
        logger.info(
            f"[{call_sid[:8]}] Caller {caller_id} free limit hit "
            f"({used_today}/{FREE_DAILY_QUERIES})"
        )
        return Response(
            twiml_say(limit_message),
            mimetype="text/xml"
        )

    # Hard cutoff: paid quota exhausted
    if used_today >= total_limit:
        logger.info(
            f"[{call_sid[:8]}] Caller {caller_id} all questions exhausted "
            f"({used_today}/{total_limit})"
        )
        return Response(
            twiml_say(
                "You've used all your questions for today. Please call back tomorrow."
            ),
            mimetype="text/xml"
        )

    logger.info(
        f"[{call_sid[:8]}] New call from {caller_id} "
        f"(used={used_today}/{total_limit}, free={FREE_DAILY_QUERIES}, paid={has_paid})"
    )

    # Show greeting + menu (ask question or manage subscription)
    resp = VoiceResponse()
    gather = Gather(
        input="dtmf",
        action="/voice-menu",
        method="POST",
        num_digits=1,
        timeout=10,
    )
    gather.say(
        SERVICE_GREETING + " "
        "Press 1 to ask a question, or 2 to manage your subscription.",
        voice=VOICE_NAME,
        language="en-US"
    )
    resp.append(gather)

    # Fallback if no selection
    resp.say("I didn't catch that. Please press 1 or 2.", voice=VOICE_NAME)
    resp.redirect("/voice")

    return Response(str(resp), mimetype="text/xml")


@app.route("/voice-menu", methods=["POST"])
def voice_menu():
    """Route to question or subscription management."""
    sig_err = enforce_twilio_signature()
    if sig_err:
        return sig_err

    call_sid = request.form.get("CallSid", "unknown")
    caller = request.form.get("From", "unknown")
    selection = request.form.get("Digits", "")

    # Store call metadata
    call_metadata[call_sid] = {"from": caller_identity(call_sid, caller)}

    if selection == "1":
        # Ask a question
        return Response(
            twiml_listen(
                "Go ahead, I'm listening.",
                action="/gather",
                fallback_text="I didn't catch that. Please ask your question.",
                redirect_url="/voice?returning=true",
            ),
            mimetype="text/xml"
        )
    elif selection == "2":
        # Manage subscription
        phone = caller.lstrip("+")
        return Response(
            twiml_subscription_menu(phone),
            mimetype="text/xml"
        )
    else:
        resp = VoiceResponse()
        resp.say("Invalid selection. Please press 1 or 2.", voice=VOICE_NAME)
        resp.redirect("/voice")
        return Response(str(resp), mimetype="text/xml")


@app.route("/intro-name", methods=["GET", "POST"])
def intro_name():
    """Capture caller name before main Q&A."""
    sig_err = enforce_twilio_signature()
    if sig_err:
        return sig_err

    call_sid = request.form.get("CallSid", "unknown")
    caller = request.form.get("From", "unknown")
    caller_id = caller_identity(call_sid, caller)
    usage_date = business_today_iso()
    spoken = request.form.get("SpeechResult", "").strip()

    if not spoken:
        return Response(
            twiml_listen(
                "I didn't catch your first name. Please say your first name now.",
                action="/intro-name",
                fallback_text="Still didn't catch that. Please say your first name.",
                redirect_url="/voice?returning=true&stage=name",
            ),
            mimetype="text/xml",
        )

    existing_name = usage_store.get_caller_name(caller_id)
    if existing_name and spoken.lower() in {"same", "same name", "use same name"}:
        caller_name = existing_name
    else:
        caller_name = extract_first_name(spoken, fallback=existing_name or "friend")
        usage_store.upsert_caller_name(caller_id, caller_name)

    if call_sid in call_metadata:
        call_metadata[call_sid]["caller_name"] = caller_name

    used_today = usage_store.get_count(caller_id, usage_date)
    has_paid = usage_store.has_paid_access(caller_id, usage_date)
    if has_paid:
        total_allowed = FREE_DAILY_QUERIES + PAID_DAILY_QUERIES
        remaining = max(0, total_allowed - used_today)
        question_word = "question" if remaining == 1 else "questions"
        quota_line = f"Paid access is active. You have {remaining} {question_word} left today."
    else:
        remaining = max(0, FREE_DAILY_QUERIES - used_today)
        question_word = "question" if remaining == 1 else "questions"
        quota_line = f"You have {remaining} free {question_word} left today."

    return Response(
        twiml_listen(
            f"Thanks, {caller_name}. {quota_line} What would you like help with first?",
            action="/gather",
            fallback_text="I'm still here. Go ahead and ask your question.",
            redirect_url="/voice?returning=true",
        ),
        mimetype="text/xml",
    )


@app.route("/gather", methods=["GET", "POST"])
def gather():
    """Handle speech input from the caller and respond via Claude."""
    sig_err = enforce_twilio_signature()
    if sig_err:
        return sig_err

    call_sid = request.form.get("CallSid", "unknown")
    caller = request.form.get("From", "unknown")
    caller_id = caller_identity(call_sid, caller)
    speech_result = request.form.get("SpeechResult", "").strip()
    usage_date = business_today_iso()

    if not speech_result:
        return Response(
            twiml_listen(
                "I didn't catch that clearly. Could you say it again?"
            ),
            mimetype="text/xml"
        )

    # Track turns
    if call_sid in call_metadata:
        call_metadata[call_sid]["turns"] += 1

    # Check for goodbye / end call
    farewell_words = [
        "goodbye", "bye", "hang up", "that's all", "thank you goodbye",
        "i'm done", "nothing else", "no thanks", "no thank you",
        "that'll do", "that will do", "all done"
    ]
    if any(word in speech_result.lower() for word in farewell_words):
        logger.info(f"[{call_sid[:8]}] Call ended by caller")
        return Response(
            twiml_say(
                "You're very welcome. Have a blessed day. Goodbye!"
            ),
            mimetype="text/xml"
        )

    has_paid = usage_store.has_paid_access(caller_id, usage_date)
    used_before = usage_store.get_count(caller_id, usage_date)
    total_limit = FREE_DAILY_QUERIES + (PAID_DAILY_QUERIES if has_paid else 0)

    if used_before >= FREE_DAILY_QUERIES and not has_paid:
        limit_message = build_limit_message(caller_id, usage_date)
        logger.info(
            f"[{call_sid[:8]}] Free limit hit for {caller_id}: "
            f"{used_before}/{FREE_DAILY_QUERIES}"
        )
        return Response(
            twiml_say(limit_message),
            mimetype="text/xml"
        )

    # Hard cutoff: paid quota exhausted
    if used_before >= total_limit:
        logger.info(
            f"[{call_sid[:8]}] All questions exhausted for {caller_id}: "
            f"{used_before}/{total_limit}"
        )
        return Response(
            twiml_say(
                "You've used all your questions for today. Please call back tomorrow."
            ),
            mimetype="text/xml"
        )

    used_after = usage_store.increment(caller_id, usage_date)
    logger.info(
        f"[{call_sid[:8]}] Counted inquiry for {caller_id}: {used_after}/{total_limit}"
    )

    policy_response = guardrail_response(speech_result)
    if policy_response:
        ai_response = policy_response
    else:
        # Get Claude's response
        try:
            ai_response = ask_claude(call_sid, speech_result)
        except Exception as e:
            logger.error(f"[{call_sid[:8]}] Claude API error: {e}")
            ai_response = (
                "I'm sorry, I'm having a little trouble thinking right now. "
                "Could you try asking me again?"
            )

    remaining = max(0, total_limit - used_after)

    # Hard cutoff: all daily questions used (free + paid quota, if unlocked)
    if remaining == 0:
        final_message = f"{ai_response} That's all your questions for today. Thanks for calling."
        return Response(twiml_say(final_message), mimetype="text/xml")

    # After final free question: require paid unlock before allowing paid quota
    if used_after == FREE_DAILY_QUERIES and not has_paid:
        limit_message = build_limit_message(caller_id, usage_date)
        message = f"{ai_response} {limit_message}"
        return Response(
            twiml_say(message),
            mimetype="text/xml"
        )

    # Continue normally
    continuation = "What else?"
    full_response = f"{ai_response} {continuation}"
    return Response(
        twiml_listen(full_response, action="/gather"),
        mimetype="text/xml"
    )


@app.route("/status", methods=["GET", "POST"])
def status():
    """Call status webhook — clean up conversation and export to spreadsheet when call ends."""
    sig_err = enforce_twilio_signature()
    if sig_err:
        return sig_err

    call_sid = request.form.get("CallSid", "")
    call_status = request.form.get("CallStatus", "")
    caller_from = request.form.get("From", "")

    if call_status in ("completed", "failed", "busy", "no-answer"):
        meta = call_metadata.pop(call_sid, {})
        history = conversations.pop(call_sid, None)
        usage_date = business_today_iso()

        if meta and call_status == "completed" and history:
            caller_id = caller_identity(call_sid, caller_from)
            caller_name = usage_store.get_caller_name(caller_id)
            has_paid = usage_store.has_paid_access(caller_id, usage_date)

            # Extract Q&A from conversation history
            questions_asked = []
            for msg in history:
                if msg["role"] == "user":
                    question = msg["content"]
                    # Find the corresponding answer
                    idx = history.index(msg)
                    answer = ""
                    if idx + 1 < len(history) and history[idx + 1]["role"] == "assistant":
                        answer = history[idx + 1]["content"]
                    questions_asked.append((question, answer))

            # Export to spreadsheet
            export_call_to_spreadsheet(caller_id, caller_name, usage_date, questions_asked, has_paid)

            logger.info(
                f"[{call_sid[:8]}] Call ended ({call_status}) — "
                f"{len(questions_asked)} Q&A, exported to spreadsheet"
            )
        elif meta:
            logger.info(
                f"[{call_sid[:8]}] Call ended ({call_status}) — "
                f"{meta.get('turns', 0)} turns"
            )

    return Response("OK", status=200)


@app.route("/square-webhook", methods=["POST"])
def square_webhook():
    """Square webhook: mark caller/day as paid when checkout completes."""
    payload = request.get_data() or b""
    if not verify_square_webhook(payload):
        return Response("invalid square signature", status=403)

    try:
        event = json.loads(payload.decode("utf-8"))
    except json.JSONDecodeError:
        return Response("invalid payload", status=400)

    event_type = (event.get("type") or "").strip().lower()
    obj = (event.get("data") or {}).get("object") or {}
    payment = obj.get("payment") if isinstance(obj, dict) else None

    if event_type in {"payment.created", "payment.updated"} and isinstance(payment, dict):
        apply_paid_event_from_square(payment)

    return Response("ok", status=200)


@app.route("/health", methods=["GET"])
def health():
    """Health check endpoint."""
    return {
        "status": "ok",
        "service": "Emmet AI",
        "phone": SERVICE_PHONE_DISPLAY,
        "active_calls": len(conversations),
        "anthropic_configured": bool(os.environ.get("ANTHROPIC_API_KEY", "").strip()),
        "model": ANTHROPIC_MODEL,
        "free_daily_queries": FREE_DAILY_QUERIES,
        "paid_daily_queries": PAID_DAILY_QUERIES,
        "business_timezone": BUSINESS_TIMEZONE,
        "square_configured": bool(SQUARE_ACCESS_TOKEN and SQUARE_LOCATION_ID),
        "voice_name": VOICE_NAME,
        "twilio_signature_validation": TWILIO_VALIDATE_SIGNATURE,
        "version": APP_VERSION,
    }


@app.route("/logs", methods=["GET"])
def logs():
    """View recent call logs (last 50 turns)."""
    date = request.args.get("date", business_today_iso())
    with usage_store._connect() as conn:
        rows = conn.execute(
            """
            SELECT call_sid, caller_phone, caller_name, turn_number,
                   user_message, assistant_message, latency_ms, created_at
            FROM call_logs
            WHERE usage_date = ?
            ORDER BY created_at DESC
            LIMIT 50
            """,
            (date,),
        ).fetchall()
    return Response(
        json.dumps([dict(r) for r in rows], indent=2),
        mimetype="application/json",
    )


ADMIN_TOKEN = env_text("ADMIN_TOKEN", "")


def require_admin(req) -> Optional[str]:
    """Return error message if request is not authorized, else None."""
    if not ADMIN_TOKEN:
        return "ADMIN_TOKEN not configured on server"
    token = req.headers.get("X-Admin-Token", "")
    if not token or token != ADMIN_TOKEN:
        return "Unauthorized"
    return None


# ── Admin: Knowledge Base ────────────────────────────────────────────────
@app.route("/admin/knowledge", methods=["GET"])
def admin_knowledge_get():
    err = require_admin(request)
    if err:
        return {"error": err}, 401
    kb = load_knowledge_base()
    return Response(json.dumps(kb, indent=2), mimetype="application/json")


@app.route("/admin/knowledge", methods=["POST"])
def admin_knowledge_post():
    """Add a Q&A entry or announcement to the live knowledge base."""
    err = require_admin(request)
    if err:
        return {"error": err}, 401
    data = request.get_json(force=True) or {}
    now = datetime.utcnow().isoformat(timespec="seconds")
    entry_type = data.get("type", "qa")  # "qa" or "announcement"
    with usage_store._connect() as conn:
        if entry_type == "announcement":
            text = data.get("announcement", "").strip()
            if not text:
                return {"error": "announcement field required"}, 400
            conn.execute(
                "INSERT INTO knowledge_updates (entry_type, announcement, active, created_at, updated_at) VALUES (?,?,1,?,?)",
                ("announcement", text, now, now),
            )
        else:
            q = data.get("question", "").strip()
            a = data.get("answer", "").strip()
            if not q or not a:
                return {"error": "question and answer fields required"}, 400
            conn.execute(
                "INSERT INTO knowledge_updates (entry_type, question, answer, active, created_at, updated_at) VALUES (?,?,?,1,?,?)",
                ("qa", q, a, now, now),
            )
        conn.commit()
    return {"status": "ok", "message": "Knowledge base updated — takes effect on next call"}


@app.route("/admin/knowledge/<int:entry_id>", methods=["DELETE"])
def admin_knowledge_delete(entry_id):
    err = require_admin(request)
    if err:
        return {"error": err}, 401
    with usage_store._connect() as conn:
        conn.execute("UPDATE knowledge_updates SET active=0 WHERE id=?", (entry_id,))
        conn.commit()
    return {"status": "ok", "message": f"Entry {entry_id} deactivated"}


# ── Admin: Marketplace ───────────────────────────────────────────────────
@app.route("/admin/marketplace", methods=["GET"])
def admin_market_get():
    err = require_admin(request)
    if err:
        return {"error": err}, 401
    cat = request.args.get("category")
    status = request.args.get("status", "available")
    query = "SELECT * FROM marketplace_listings"
    params = []
    filters = []
    if cat:
        filters.append("category=?")
        params.append(cat.lower())
    if status != "all":
        filters.append("status=?")
        params.append(status)
    if filters:
        query += " WHERE " + " AND ".join(filters)
    query += " ORDER BY category, id DESC"
    with usage_store._connect() as conn:
        rows = conn.execute(query, params).fetchall()
    return Response(json.dumps([dict(r) for r in rows], indent=2), mimetype="application/json")


@app.route("/admin/marketplace", methods=["POST"])
def admin_market_post():
    """Add a new marketplace listing."""
    err = require_admin(request)
    if err:
        return {"error": err}, 401
    data = request.get_json(force=True) or {}
    required = ["category", "title", "description"]
    missing = [f for f in required if not data.get(f, "").strip()]
    if missing:
        return {"error": f"Missing required fields: {missing}"}, 400
    now = datetime.utcnow().isoformat(timespec="seconds")
    with usage_store._connect() as conn:
        cur = conn.execute(
            """INSERT INTO marketplace_listings
               (category, title, description, price, location, contact, status, listed_by, created_at, updated_at)
               VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (
                data["category"].lower().strip(),
                data["title"].strip(),
                data["description"].strip(),
                data.get("price", ""),
                data.get("location", ""),
                data.get("contact", ""),
                data.get("status", "available"),
                data.get("listed_by", ""),
                now, now,
            ),
        )
        new_id = cur.lastrowid
        conn.commit()
    return {"status": "ok", "id": new_id, "message": "Listing added — callers can ask about it immediately"}


@app.route("/admin/marketplace/<int:listing_id>", methods=["PATCH"])
def admin_market_patch(listing_id):
    """Update a listing — e.g. mark as sold."""
    err = require_admin(request)
    if err:
        return {"error": err}, 401
    data = request.get_json(force=True) or {}
    allowed = ["title", "description", "price", "location", "contact", "status", "listed_by"]
    updates = {k: v for k, v in data.items() if k in allowed}
    if not updates:
        return {"error": "No valid fields to update"}, 400
    now = datetime.utcnow().isoformat(timespec="seconds")
    updates["updated_at"] = now
    set_clause = ", ".join(f"{k}=?" for k in updates)
    values = list(updates.values()) + [listing_id]
    with usage_store._connect() as conn:
        conn.execute(f"UPDATE marketplace_listings SET {set_clause} WHERE id=?", values)
        conn.commit()
    return {"status": "ok", "message": f"Listing {listing_id} updated"}


@app.route("/admin/marketplace/<int:listing_id>", methods=["DELETE"])
def admin_market_delete(listing_id):
    err = require_admin(request)
    if err:
        return {"error": err}, 401
    with usage_store._connect() as conn:
        conn.execute("DELETE FROM marketplace_listings WHERE id=?", (listing_id,))
        conn.commit()
    return {"status": "ok", "message": f"Listing {listing_id} removed"}


# ── Admin Dashboard & API ─────────────────────────────────────────────────

@app.route("/admin/dashboard")
def admin_dashboard():
    """Serve the admin dashboard HTML."""
    dash_path = os.path.join(os.path.dirname(__file__), "admin.html")
    if os.path.exists(dash_path):
        with open(dash_path, "r") as f:
            return Response(f.read(), mimetype="text/html")
    return "Dashboard not found", 404


@app.route("/admin/stats", methods=["GET"])
def admin_stats():
    """Overview statistics for the dashboard."""
    err = require_admin(request)
    if err:
        return {"error": err}, 401
    try:
        with usage_store._connect() as conn:
            conn.row_factory = sqlite3.Row
            total_customers = conn.execute("SELECT COUNT(*) as c FROM customers").fetchone()["c"]
            active_subs = conn.execute("SELECT COUNT(*) as c FROM subscriptions WHERE status='active'").fetchone()["c"]
            total_revenue = conn.execute("SELECT COALESCE(SUM(amount_cents),0) as c FROM payments WHERE status='completed'").fetchone()["c"]
            today = datetime.now(ZoneInfo(BUSINESS_TIMEZONE)).strftime("%Y-%m-%d")
            calls_today = conn.execute("SELECT COUNT(DISTINCT call_sid) as c FROM call_logs WHERE usage_date=?", (today,)).fetchone()["c"]
            recent_payments = [dict(r) for r in conn.execute(
                "SELECT p.*, c.phone FROM payments p LEFT JOIN customers c ON p.customer_id=c.id ORDER BY p.created_at DESC LIMIT 5"
            ).fetchall()]
            recent_calls = [dict(r) for r in conn.execute(
                "SELECT call_sid, caller_phone, caller_name, MAX(turn_number) as turns, usage_date, MAX(created_at) as last_at FROM call_logs GROUP BY call_sid ORDER BY last_at DESC LIMIT 5"
            ).fetchall()]
        return {
            "total_customers": total_customers,
            "active_subscriptions": active_subs,
            "total_revenue_cents": total_revenue,
            "calls_today": calls_today,
            "recent_payments": recent_payments,
            "recent_calls": recent_calls,
        }
    except Exception as e:
        logger.error(f"Admin stats error: {e}")
        return {"error": str(e)}, 500


@app.route("/admin/customers", methods=["GET"])
def admin_customers():
    """List all customers with optional search."""
    err = require_admin(request)
    if err:
        return {"error": err}, 401
    search = request.args.get("q", "").strip()
    try:
        with usage_store._connect() as conn:
            conn.row_factory = sqlite3.Row
            if search:
                rows = conn.execute(
                    "SELECT * FROM customers WHERE phone LIKE ? OR first_name LIKE ? OR last_name LIKE ? ORDER BY created_at DESC",
                    (f"%{search}%", f"%{search}%", f"%{search}%")
                ).fetchall()
            else:
                rows = conn.execute("SELECT * FROM customers ORDER BY created_at DESC").fetchall()
        return Response(json.dumps([dict(r) for r in rows], indent=2), mimetype="application/json")
    except Exception as e:
        return {"error": str(e)}, 500


@app.route("/admin/subscriptions", methods=["GET"])
def admin_subscriptions():
    """List all subscriptions with customer info."""
    err = require_admin(request)
    if err:
        return {"error": err}, 401
    plan = request.args.get("plan", "").strip()
    status = request.args.get("status", "").strip()
    try:
        with usage_store._connect() as conn:
            conn.row_factory = sqlite3.Row
            query = """SELECT s.*, c.phone, c.first_name, c.last_name
                       FROM subscriptions s
                       LEFT JOIN customers c ON s.customer_id=c.id
                       WHERE 1=1"""
            params = []
            if plan:
                query += " AND s.plan_tier=?"
                params.append(plan)
            if status:
                query += " AND s.status=?"
                params.append(status)
            query += " ORDER BY s.created_at DESC"
            rows = conn.execute(query, params).fetchall()
        return Response(json.dumps([dict(r) for r in rows], indent=2), mimetype="application/json")
    except Exception as e:
        return {"error": str(e)}, 500


@app.route("/admin/payments", methods=["GET"])
def admin_payments():
    """List all payments with customer info."""
    err = require_admin(request)
    if err:
        return {"error": err}, 401
    status = request.args.get("status", "").strip()
    try:
        with usage_store._connect() as conn:
            conn.row_factory = sqlite3.Row
            query = """SELECT p.*, c.phone, c.first_name, c.last_name
                       FROM payments p
                       LEFT JOIN customers c ON p.customer_id=c.id
                       WHERE 1=1"""
            params = []
            if status:
                query += " AND p.status=?"
                params.append(status)
            query += " ORDER BY p.created_at DESC"
            rows = conn.execute(query, params).fetchall()
        return Response(json.dumps([dict(r) for r in rows], indent=2), mimetype="application/json")
    except Exception as e:
        return {"error": str(e)}, 500


@app.route("/admin/call-logs", methods=["GET"])
def admin_call_logs():
    """List recent call logs."""
    err = require_admin(request)
    if err:
        return {"error": err}, 401
    limit = int(request.args.get("limit", 100))
    try:
        with usage_store._connect() as conn:
            conn.row_factory = sqlite3.Row
            rows = conn.execute(
                "SELECT call_sid, caller_phone, caller_name, turn_number, user_message, assistant_message, latency_ms, usage_date, created_at FROM call_logs ORDER BY created_at DESC LIMIT ?", (limit,)
            ).fetchall()
        return Response(json.dumps([dict(r) for r in rows], indent=2), mimetype="application/json")
    except Exception as e:
        return {"error": str(e)}, 500


# ── Unified Service Hub (SSO Dashboard APIs) ─────────────────────────────

@app.route("/admin/services/twilio", methods=["GET"])
def admin_twilio_status():
    """Pull live Twilio account data: balance, calls, messages, phone numbers."""
    err = require_admin(request)
    if err:
        return {"error": err}, 401
    if not TWILIO_ACCOUNT_SID or not TWILIO_AUTH_TOKEN:
        return {"error": "Twilio not configured", "connected": False}, 200
    try:
        from twilio.rest import Client as TC
        tc = TC(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)
        # Account info
        acct = tc.api.accounts(TWILIO_ACCOUNT_SID).fetch()
        # Balance
        bal_list = tc.api.accounts(TWILIO_ACCOUNT_SID).balance.fetch()
        balance = bal_list.balance if bal_list else "N/A"
        currency = bal_list.currency if bal_list else "USD"
        # Phone numbers
        numbers = [{"phone": n.phone_number, "friendly": n.friendly_name, "capabilities": n.capabilities}
                   for n in tc.incoming_phone_numbers.list(limit=10)]
        # Recent calls (last 20)
        recent_calls = []
        for c in tc.calls.list(limit=20):
            recent_calls.append({
                "sid": c.sid, "from": c.from_formatted, "to": c.to_formatted,
                "status": c.status, "direction": c.direction,
                "duration": c.duration, "date": str(c.date_created),
            })
        # Recent messages (last 20)
        recent_msgs = []
        for m in tc.messages.list(limit=20):
            recent_msgs.append({
                "sid": m.sid, "from": m.from_, "to": m.to,
                "status": m.status, "direction": m.direction,
                "body": m.body[:100] if m.body else "", "date": str(m.date_sent),
            })
        return {
            "connected": True,
            "account_name": acct.friendly_name,
            "account_sid": TWILIO_ACCOUNT_SID[:8] + "...",
            "status": acct.status,
            "balance": balance,
            "currency": currency,
            "phone_numbers": numbers,
            "recent_calls": recent_calls,
            "recent_messages": recent_msgs,
            "console_url": "https://console.twilio.com",
        }
    except Exception as e:
        logger.error(f"Twilio status error: {e}")
        return {"connected": False, "error": str(e)}, 200


@app.route("/admin/services/square", methods=["GET"])
def admin_square_status():
    """Pull live Square account data: locations, transactions, balance."""
    err = require_admin(request)
    if err:
        return {"error": err}, 401
    if not SQUARE_ACCESS_TOKEN:
        return {"error": "Square not configured", "connected": False}, 200
    try:
        base = "https://connect.squareup.com" if SQUARE_ENVIRONMENT == "production" else "https://connect.squareupsandbox.com"
        headers = {
            "Authorization": f"Bearer {SQUARE_ACCESS_TOKEN}",
            "Content-Type": "application/json",
            "Square-Version": SQUARE_API_VERSION,
        }
        # Locations
        loc_resp = requests.get(f"{base}/v2/locations", headers=headers, timeout=10)
        locations = []
        if loc_resp.ok:
            for loc in loc_resp.json().get("locations", []):
                locations.append({
                    "id": loc.get("id"), "name": loc.get("name"),
                    "status": loc.get("status"), "currency": loc.get("currency"),
                    "address": loc.get("address", {}).get("address_line_1", ""),
                })
        # Recent payments
        pay_resp = requests.get(f"{base}/v2/payments?limit=20&sort_order=DESC", headers=headers, timeout=10)
        payments = []
        if pay_resp.ok:
            for p in pay_resp.json().get("payments", []):
                amt = p.get("amount_money", {})
                payments.append({
                    "id": p.get("id"), "status": p.get("status"),
                    "amount_cents": amt.get("amount", 0),
                    "currency": amt.get("currency", "USD"),
                    "source_type": p.get("source_type", ""),
                    "created_at": p.get("created_at", ""),
                })
        # Merchant info
        merchant_resp = requests.get(f"{base}/v2/merchants/me", headers=headers, timeout=10)
        merchant = {}
        if merchant_resp.ok:
            m = merchant_resp.json().get("merchant", {})
            merchant = {"business_name": m.get("business_name"), "country": m.get("country"), "currency": m.get("currency")}
        return {
            "connected": True,
            "environment": SQUARE_ENVIRONMENT,
            "merchant": merchant,
            "locations": locations,
            "recent_payments": payments,
            "dashboard_url": "https://squareup.com/dashboard" if SQUARE_ENVIRONMENT == "production" else "https://squareupsandbox.com/dashboard",
        }
    except Exception as e:
        logger.error(f"Square status error: {e}")
        return {"connected": False, "error": str(e)}, 200


@app.route("/admin/services/github", methods=["GET"])
def admin_github_status():
    """Pull live GitHub repo data: recent commits, issues, deploy status."""
    err = require_admin(request)
    if err:
        return {"error": err}, 401
    repo = "josephkunin-cmyk/emmetai-agent"
    gh_headers = {"Accept": "application/vnd.github.v3+json"}
    try:
        # Repo info
        repo_resp = requests.get(f"https://api.github.com/repos/{repo}", headers=gh_headers, timeout=10)
        repo_info = {}
        if repo_resp.ok:
            r = repo_resp.json()
            repo_info = {
                "name": r.get("full_name"), "private": r.get("private"),
                "default_branch": r.get("default_branch"),
                "updated_at": r.get("updated_at"), "pushed_at": r.get("pushed_at"),
                "size_kb": r.get("size"), "open_issues": r.get("open_issues_count"),
            }
        # Recent commits
        commits_resp = requests.get(f"https://api.github.com/repos/{repo}/commits?per_page=10", headers=gh_headers, timeout=10)
        commits = []
        if commits_resp.ok:
            for c in commits_resp.json():
                commits.append({
                    "sha": c.get("sha", "")[:7],
                    "message": c.get("commit", {}).get("message", "").split("\n")[0][:80],
                    "author": c.get("commit", {}).get("author", {}).get("name", ""),
                    "date": c.get("commit", {}).get("author", {}).get("date", ""),
                })
        return {
            "connected": True,
            "repo": repo_info,
            "recent_commits": commits,
            "repo_url": f"https://github.com/{repo}",
        }
    except Exception as e:
        logger.error(f"GitHub status error: {e}")
        return {"connected": False, "error": str(e)}, 200


@app.route("/admin/services/render", methods=["GET"])
def admin_render_status():
    """Show Render deployment info based on what we know."""
    err = require_admin(request)
    if err:
        return {"error": err}, 401
    return {
        "connected": True,
        "service_name": "emmetai-agent",
        "service_url": PUBLIC_BASE_URL,
        "dashboard_url": "https://dashboard.render.com",
        "environment": {
            "ANTHROPIC_MODEL": ANTHROPIC_MODEL,
            "VOICE_NAME": VOICE_NAME,
            "FREE_DAILY_QUERIES": FREE_DAILY_QUERIES,
            "PAID_DAILY_QUERIES": PAID_DAILY_QUERIES,
            "SQUARE_ENVIRONMENT": SQUARE_ENVIRONMENT,
            "TWILIO_CONFIGURED": bool(TWILIO_ACCOUNT_SID),
            "SQUARE_CONFIGURED": bool(SQUARE_ACCESS_TOKEN),
            "ADMIN_TOKEN_SET": bool(ADMIN_TOKEN),
        },
    }


@app.route("/admin/services/all", methods=["GET"])
def admin_all_services():
    """Single endpoint to fetch status from all services at once."""
    err = require_admin(request)
    if err:
        return {"error": err}, 401
    results = {}
    # Twilio
    try:
        results["twilio"] = admin_twilio_status().get_json() if hasattr(admin_twilio_status(), 'get_json') else json.loads(admin_twilio_status().data) if hasattr(admin_twilio_status(), 'data') else {"connected": False}
    except Exception:
        pass
    # Build from individual calls
    with app.test_request_context(headers={"X-Admin-Token": request.headers.get("X-Admin-Token", "")}):
        try:
            tw = admin_twilio_status()
            results["twilio"] = tw.get_json() if hasattr(tw, 'get_json') else json.loads(tw[0]) if isinstance(tw, tuple) else {"connected": False}
        except Exception as e:
            results["twilio"] = {"connected": False, "error": str(e)}
        try:
            sq = admin_square_status()
            results["square"] = sq.get_json() if hasattr(sq, 'get_json') else json.loads(sq[0]) if isinstance(sq, tuple) else {"connected": False}
        except Exception as e:
            results["square"] = {"connected": False, "error": str(e)}
        try:
            gh = admin_github_status()
            results["github"] = gh.get_json() if hasattr(gh, 'get_json') else json.loads(gh[0]) if isinstance(gh, tuple) else {"connected": False}
        except Exception as e:
            results["github"] = {"connected": False, "error": str(e)}
        try:
            rn = admin_render_status()
            results["render"] = rn.get_json() if hasattr(rn, 'get_json') else json.loads(rn[0]) if isinstance(rn, tuple) else {"connected": False}
        except Exception as e:
            results["render"] = {"connected": False, "error": str(e)}
    return results


# ── Voice Payment System (Square Integration) ────────────────────────────
# Global dict to store in-flight payment sessions (card details temporary)
payment_sessions = {}

def initialize_customer_database():
    """Create customer & payment tables if they don't exist."""
    try:
        with usage_store._connect() as conn:
            # Customers table
            conn.execute("""
                CREATE TABLE IF NOT EXISTS customers (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    phone TEXT UNIQUE NOT NULL,
                    email TEXT,
                    first_name TEXT,
                    last_name TEXT,
                    external_id TEXT UNIQUE,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    last_call_at TEXT,
                    status TEXT DEFAULT 'active'
                )
            """)

            # Subscriptions table
            conn.execute("""
                CREATE TABLE IF NOT EXISTS subscriptions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    customer_id INTEGER NOT NULL UNIQUE REFERENCES customers(id),
                    plan_tier TEXT NOT NULL,
                    plan_name TEXT,
                    status TEXT DEFAULT 'active',
                    started_at TEXT NOT NULL,
                    renewal_date TEXT,
                    canceled_at TEXT,
                    cancel_reason TEXT,
                    auto_renew INTEGER DEFAULT 1,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                )
            """)

            # Payments table
            conn.execute("""
                CREATE TABLE IF NOT EXISTS payments (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    customer_id INTEGER NOT NULL REFERENCES customers(id),
                    subscription_id INTEGER REFERENCES subscriptions(id),
                    amount_cents INTEGER NOT NULL,
                    currency TEXT DEFAULT 'USD',
                    order_id TEXT UNIQUE,
                    payment_link_id TEXT UNIQUE,
                    payment_link_url TEXT,
                    source TEXT DEFAULT 'phone',
                    status TEXT DEFAULT 'pending',
                    paid_at TEXT,
                    failed_at TEXT,
                    failure_reason TEXT,
                    refund_amount_cents INTEGER DEFAULT 0,
                    refunded_at TEXT,
                    refund_reason TEXT,
                    retry_count INTEGER DEFAULT 0,
                    next_retry_at TEXT,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                )
            """)

            conn.commit()
            logger.info("✅ Customer database tables initialized")
    except Exception as e:
        logger.error(f"Error initializing customer DB: {e}")

def get_or_create_customer(phone: str) -> dict:
    """Get existing customer or create new one."""
    try:
        with usage_store._connect() as conn:
            row = conn.execute(
                "SELECT * FROM customers WHERE phone = ?",
                (phone,)
            ).fetchone()

            if row:
                return dict(row)

            # Create new customer
            now = datetime.utcnow().isoformat()
            conn.execute(
                "INSERT INTO customers (phone, created_at, updated_at) VALUES (?, ?, ?)",
                (phone, now, now)
            )
            conn.commit()

            row = conn.execute(
                "SELECT * FROM customers WHERE phone = ?",
                (phone,)
            ).fetchone()
            return dict(row) if row else {}
    except Exception as e:
        logger.error(f"Error getting/creating customer: {e}")
        return {}

def validate_card_number(card_num: str) -> bool:
    """Validate card with Luhn algorithm."""
    if not card_num.isdigit() or len(card_num) < 13 or len(card_num) > 19:
        return False
    total = 0
    for i, digit in enumerate(card_num[::-1]):
        n = int(digit)
        if i % 2 == 1:
            n *= 2
            if n > 9:
                n -= 9
        total += n
    return total % 10 == 0

def is_valid_expiration(month: str, year: str) -> bool:
    """Check if card hasn't expired."""
    try:
        current = datetime.utcnow()
        exp_date = datetime(int(year), int(month), 1)
        exp_date = exp_date.replace(day=1) + timedelta(days=32)
        return exp_date > current
    except:
        return False

def detect_card_type(card_num: str) -> str:
    """Detect card type from card number."""
    if card_num.startswith("4"):
        return "Visa"
    elif card_num.startswith("5"):
        return "Mastercard"
    elif card_num.startswith("3"):
        return "Amex"
    elif card_num.startswith("6"):
        return "Discover"
    else:
        return "Unknown"

def charge_card_square(card_number: str, exp_month: str, exp_year: str,
                      cvv: str, amount_cents: int, customer_phone: str, plan: str) -> dict:
    """Charge card via Square Payment Intents API."""
    try:
        import requests

        headers = {
            "Authorization": f"Bearer {SQUARE_ACCESS_TOKEN}",
            "Content-Type": "application/json",
            "Square-Version": SQUARE_API_VERSION,
        }

        # Create/get customer
        customer = get_or_create_customer(customer_phone)
        if not customer.get("id"):
            return {"success": False, "error": "Could not create customer"}

        # Create payment
        payment_body = {
            "idempotency_key": str(uuid.uuid4()),
            "amount_money": {
                "amount": amount_cents,
                "currency": SQUARE_CURRENCY,
            },
            "payment_source_details": {
                "card_details": {
                    "card": {
                        "card_number": card_number,
                        "expiration_month": int(exp_month),
                        "expiration_year": int(exp_year),
                        "cvc": cvv,
                    }
                }
            },
            "receipt_number": f"emmet-{customer_phone}-{int(time.time())}",
            "note": f"Subscription: {plan}",
        }

        payment_response = requests.post(
            "https://connect.squareup.com/v2/payments",
            json=payment_body,
            headers=headers,
            timeout=10,
        ).json()

        if "error" in payment_response:
            error_msg = payment_response["error"].get("message", "Unknown error")
            logger.error(f"Square payment error: {error_msg}")
            return {"success": False, "error": error_msg}

        payment_id = payment_response["payment"]["id"]

        # Create subscription in DB
        now = datetime.utcnow().isoformat()
        renewal_date = (datetime.utcnow() + timedelta(days=30)).isoformat()

        plans = {
            "paid_9": ("Basic", 999),
            "paid_29": ("Professional", 2999),
            "paid_59": ("Enterprise", 5999),
        }

        plan_name, _ = plans.get(plan, ("Premium", amount_cents))

        with usage_store._connect() as conn:
            conn.execute(
                """
                INSERT OR REPLACE INTO subscriptions
                (customer_id, plan_tier, plan_name, status, started_at,
                 renewal_date, created_at, updated_at)
                VALUES (?, ?, ?, 'active', ?, ?, ?, ?)
                """,
                (customer["id"], plan, plan_name, now, renewal_date, now, now)
            )

            conn.execute(
                """
                INSERT INTO payments
                (customer_id, amount_cents, order_id, status, paid_at, created_at, updated_at)
                VALUES (?, ?, ?, 'completed', ?, ?, ?)
                """,
                (customer["id"], amount_cents, payment_id, now, now, now)
            )

            conn.commit()

        # Send SMS confirmation
        send_sms(
            customer_phone,
            f"✅ Subscription confirmed! Charged ${amount_cents / 100:.2f} for "
            f"{plan_name} plan. Next billing: {renewal_date[:10]}. "
            f"Call (717) 922-5968 to manage."
        )

        return {
            "success": True,
            "payment_id": payment_id,
            "plan_name": plan_name,
            "amount": amount_cents / 100,
        }

    except Exception as e:
        logger.error(f"Error charging card: {e}")
        return {"success": False, "error": str(e)}

# Voice Payment TwiML Endpoints

def twiml_subscription_menu(phone: str) -> str:
    """Show subscription plan options."""
    resp = VoiceResponse()
    gather = Gather(
        input="dtmf",
        action=f"/gather-plan-selection?phone={phone}",
        method="POST",
        num_digits=1,
        timeout=10,
    )
    gather.say(
        "Choose a plan. Press 1 for Basic: four questions daily for nine dollars ninety-nine per month. "
        "Press 2 for Professional: unlimited questions for twenty-nine dollars ninety-nine per month. "
        "Press 3 for Enterprise: unlimited with priority for fifty-nine dollars ninety-nine per month. "
        "Press 0 to go back.",
        voice=VOICE_NAME,
        language="en-US"
    )
    resp.append(gather)
    resp.say("Sorry, I didn't catch that.", voice=VOICE_NAME)
    resp.redirect("/subscribe-menu?phone=" + phone)
    return str(resp)

@app.route("/subscribe-menu", methods=["GET", "POST"])
def subscribe_menu():
    """Show subscription menu."""
    sig_err = enforce_twilio_signature()
    if sig_err:
        return sig_err

    phone = request.args.get("phone", request.form.get("From", "unknown"))
    call_sid = request.form.get("CallSid", "unknown")
    call_metadata[call_sid] = {"from": phone}
    return Response(twiml_subscription_menu(phone), mimetype="text/xml")

@app.route("/gather-plan-selection", methods=["POST"])
def gather_plan_selection():
    """Handle plan selection."""
    sig_err = enforce_twilio_signature()
    if sig_err:
        return sig_err

    phone = request.args.get("phone", "unknown")
    selection = request.form.get("Digits", "")

    plans = {
        "1": {"tier": "paid_9", "name": "Basic", "price_cents": 999},
        "2": {"tier": "paid_29", "name": "Professional", "price_cents": 2999},
        "3": {"tier": "paid_59", "name": "Enterprise", "price_cents": 5999},
    }

    if selection not in plans:
        return Response(
            twiml_say("Invalid selection. Please try again.") +
            twiml_listen("Press 1, 2, or 3 to select a plan.", action="/subscribe-menu?phone=" + phone),
            mimetype="text/xml"
        )

    plan = plans[selection]

    resp = VoiceResponse()
    gather = Gather(
        input="dtmf",
        action=f"/gather-payment-method?phone={phone}&plan={plan['tier']}&amount={plan['price_cents']}",
        method="POST",
        num_digits=1,
        timeout=10,
    )
    gather.say(
        f"You selected {plan['name']} for ${plan['price_cents']/100:.2f} per month. "
        "How would you like to pay? Press 1 for credit card over the phone, "
        "or press 2 to receive a payment link via text.",
        voice=VOICE_NAME,
        language="en-US"
    )
    resp.append(gather)
    resp.say("I didn't catch that. Please try again.", voice=VOICE_NAME)
    resp.redirect(f"/gather-plan-selection?phone={phone}")
    return Response(str(resp), mimetype="text/xml")

@app.route("/gather-payment-method", methods=["POST"])
def gather_payment_method():
    """Route to voice or SMS payment."""
    sig_err = enforce_twilio_signature()
    if sig_err:
        return sig_err

    phone = request.args.get("phone", "unknown")
    plan = request.args.get("plan", "paid_9")
    amount = int(request.args.get("amount", 999))
    selection = request.form.get("Digits", "")

    if selection == "1":
        # Collect card number
        resp = VoiceResponse()
        gather = Gather(
            input="dtmf",
            action=f"/process-card-number?phone={phone}&plan={plan}&amount={amount}",
            method="POST",
            num_digits=19,
            timeout=15,
            finish_on_key="#",
        )
        gather.say(
            "Please enter your credit card number, then press the pound key.",
            voice=VOICE_NAME,
            language="en-US"
        )
        resp.append(gather)
        resp.say("I didn't receive a card number. Please try again.", voice=VOICE_NAME)
        resp.redirect(f"/gather-payment-method?phone={phone}&plan={plan}&amount={amount}")
        return Response(str(resp), mimetype="text/xml")

    elif selection == "2":
        # Send SMS link
        return Response(
            twiml_send_payment_link(phone, plan, amount),
            mimetype="text/xml"
        )
    else:
        resp = VoiceResponse()
        resp.say("Invalid selection. Please try again.", voice=VOICE_NAME)
        resp.redirect(f"/gather-plan-selection?phone={phone}")
        return Response(str(resp), mimetype="text/xml")

@app.route("/process-card-number", methods=["POST"])
def process_card_number():
    """Process card number, ask for expiration."""
    sig_err = enforce_twilio_signature()
    if sig_err:
        return sig_err

    phone = request.args.get("phone", "unknown")
    plan = request.args.get("plan", "paid_9")
    amount = int(request.args.get("amount", 999))
    card_number = request.form.get("Digits", "").replace("#", "").strip()

    if not validate_card_number(card_number):
        resp = VoiceResponse()
        resp.say(
            "That doesn't look like a valid card number. Please check and try again.",
            voice=VOICE_NAME
        )
        resp.redirect(f"/gather-payment-method?phone={phone}&plan={plan}&amount={amount}")
        return Response(str(resp), mimetype="text/xml")

    session_key = f"payment_{phone}_{int(time.time())}"
    payment_sessions[session_key] = {
        "phone": phone,
        "plan": plan,
        "amount": amount,
        "card_number": card_number,
        "card_last4": card_number[-4:],
        "card_type": detect_card_type(card_number),
        "created_at": time.time(),
    }

    resp = VoiceResponse()
    gather = Gather(
        input="dtmf",
        action=f"/process-card-expiration?session={session_key}",
        method="POST",
        num_digits=4,
        timeout=10,
        finish_on_key="#",
    )
    gather.say(
        f"Thank you. Card ending in {card_number[-4:]}. "
        "Now enter the expiration month and year (M M Y Y), then press pound. "
        "For example, for December 2027, enter 1 2 2 7 and press pound.",
        voice=VOICE_NAME,
        language="en-US"
    )
    resp.append(gather)
    resp.say("I didn't receive the expiration date. Please try again.", voice=VOICE_NAME)
    resp.redirect(f"/gather-payment-method?phone={phone}&plan={plan}&amount={amount}")
    return Response(str(resp), mimetype="text/xml")

@app.route("/process-card-expiration", methods=["POST"])
def process_card_expiration():
    """Process expiration, ask for CVV."""
    sig_err = enforce_twilio_signature()
    if sig_err:
        return sig_err

    session_key = request.args.get("session")
    expiration = request.form.get("Digits", "").replace("#", "").strip()

    if session_key not in payment_sessions:
        resp = VoiceResponse()
        resp.say("Your session expired. Please call back and try again.", voice=VOICE_NAME)
        resp.hangup()
        return Response(str(resp), mimetype="text/xml")

    session = payment_sessions[session_key]

    if len(expiration) != 4 or not expiration.isdigit():
        resp = VoiceResponse()
        resp.say("That doesn't look valid. Please enter M M Y Y.", voice=VOICE_NAME)
        resp.redirect(f"/gather-payment-method?phone={session['phone']}&plan={session['plan']}&amount={session['amount']}")
        return Response(str(resp), mimetype="text/xml")

    month = expiration[:2]
    year = "20" + expiration[2:]

    if not is_valid_expiration(month, year):
        resp = VoiceResponse()
        resp.say("That expiration date has passed. Please check your card and try again.", voice=VOICE_NAME)
        resp.redirect(f"/gather-payment-method?phone={session['phone']}&plan={session['plan']}&amount={session['amount']}")
        return Response(str(resp), mimetype="text/xml")

    payment_sessions[session_key]["exp_month"] = month
    payment_sessions[session_key]["exp_year"] = year

    resp = VoiceResponse()
    gather = Gather(
        input="dtmf",
        action=f"/process-card-cvv?session={session_key}",
        method="POST",
        num_digits=4,
        timeout=10,
        finish_on_key="#",
    )
    gather.say(
        f"Expiration: {month}/{year}. "
        "Now enter the 3-digit security code on the back of your card, then press pound.",
        voice=VOICE_NAME,
        language="en-US"
    )
    resp.append(gather)
    resp.say("I didn't receive the security code. Please try again.", voice=VOICE_NAME)
    resp.redirect(f"/gather-payment-method?phone={session['phone']}&plan={session['plan']}&amount={session['amount']}")
    return Response(str(resp), mimetype="text/xml")

@app.route("/process-card-cvv", methods=["POST"])
def process_card_cvv():
    """Charge card via Square."""
    sig_err = enforce_twilio_signature()
    if sig_err:
        return sig_err

    session_key = request.args.get("session")
    cvv = request.form.get("Digits", "").replace("#", "").strip()

    if session_key not in payment_sessions:
        resp = VoiceResponse()
        resp.say("Your session expired. Please call back.", voice=VOICE_NAME)
        resp.hangup()
        return Response(str(resp), mimetype="text/xml")

    session = payment_sessions[session_key]

    if len(cvv) < 3 or len(cvv) > 4 or not cvv.isdigit():
        resp = VoiceResponse()
        resp.say("That doesn't look like a valid security code. Please call back to try again.", voice=VOICE_NAME)
        resp.hangup()
        return Response(str(resp), mimetype="text/xml")

    # Charge card
    charge_result = charge_card_square(
        card_number=session["card_number"],
        exp_month=session["exp_month"],
        exp_year=session["exp_year"],
        cvv=cvv,
        amount_cents=session["amount"],
        customer_phone=session["phone"],
        plan=session["plan"]
    )

    # Clean up session
    if session_key in payment_sessions:
        del payment_sessions[session_key]

    if not charge_result["success"]:
        resp = VoiceResponse()
        resp.say(
            f"Unfortunately, your card was declined. Your card has not been charged. "
            "I can send you a payment link via text to try a different card. "
            "To receive a link, press 1. Otherwise, just hang up and call back later.",
            voice=VOICE_NAME
        )
        gather = Gather(
            input="dtmf",
            action=f"/fallback-sms-link?phone={session['phone']}&plan={session['plan']}&amount={session['amount']}",
            method="POST",
            num_digits=1,
            timeout=10,
        )
        gather.say("Press 1 for SMS link, or hang up.", voice=VOICE_NAME)
        resp.append(gather)
        resp.hangup()
        return Response(str(resp), mimetype="text/xml")

    # SUCCESS
    resp = VoiceResponse()
    resp.say(
        f"Success! Your card ending in {session['card_last4']} has been charged "
        f"${session['amount'] / 100:.2f}. Your subscription is now active. "
        "You will be automatically charged on the same date each month. "
        "A confirmation text has been sent to your phone. Thank you!",
        voice=VOICE_NAME
    )
    resp.hangup()
    return Response(str(resp), mimetype="text/xml")

def twiml_send_payment_link(phone: str, plan: str, amount: int) -> str:
    """Send Square payment link via SMS."""
    try:
        import requests

        plans_info = {
            "paid_9": {"name": "Basic - $9.99/month", "amount": 999},
            "paid_29": {"name": "Professional - $29.99/month", "amount": 2999},
            "paid_59": {"name": "Enterprise - $59.99/month", "amount": 5999},
        }

        plan_details = plans_info.get(plan, {"name": "Subscription", "amount": amount})

        headers = {
            "Authorization": f"Bearer {SQUARE_ACCESS_TOKEN}",
            "Content-Type": "application/json",
            "Square-Version": SQUARE_API_VERSION,
        }

        link_body = {
            "idempotency_key": str(uuid.uuid4()),
            "quick_pay": {
                "name": plan_details["name"],
                "price_money": {
                    "amount": plan_details["amount"],
                    "currency": SQUARE_CURRENCY,
                },
            },
        }

        link_response = requests.post(
            "https://connect.squareup.com/v2/checkout/payment-links",
            json=link_body,
            headers=headers,
            timeout=10,
        ).json()

        payment_link_url = link_response.get("payment_link", {}).get("url")

        if not payment_link_url:
            resp = VoiceResponse()
            resp.say("I'm having trouble generating a link. Visit emmetai.com to subscribe.", voice=VOICE_NAME)
            resp.hangup()
            return str(resp)

        send_sms(
            phone,
            f"Pay for {plan_details['name']}: {payment_link_url} (Link expires in 30 min)"
        )

        resp = VoiceResponse()
        resp.say(
            f"Perfect! I've sent you a text with a payment link for {plan_details['name']}. "
            "Click the link to finish paying. Thank you!",
            voice=VOICE_NAME
        )
        resp.hangup()
        return str(resp)

    except Exception as e:
        logger.error(f"Error sending SMS link: {e}")
        resp = VoiceResponse()
        resp.say("I'm having trouble. Please try again later or visit emmetai.com.", voice=VOICE_NAME)
        resp.hangup()
        return str(resp)

@app.route("/fallback-sms-link", methods=["POST"])
def fallback_sms_link():
    """Fallback to SMS if voice payment fails."""
    sig_err = enforce_twilio_signature()
    if sig_err:
        return sig_err

    phone = request.args.get("phone", "unknown")
    plan = request.args.get("plan", "paid_9")
    amount = int(request.args.get("amount", 999))
    return Response(
        twiml_send_payment_link(phone, plan, amount),
        mimetype="text/xml"
    )

# Initialize databases on startup
initialize_farming_knowledge_db()
initialize_customer_database()


# ── Entry Point ──────────────────────────────────────────────────────────
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    logger.info(f"Starting Emmet AI agent on port {port}")
    app.run(host="0.0.0.0", port=port, debug=False)
# Updated Mon Mar  2 00:38:10 UTC 2026
